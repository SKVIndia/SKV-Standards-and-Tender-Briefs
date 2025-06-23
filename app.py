import streamlit as st
import pandas as pd
import numpy as np
from sentence_transformers import SentenceTransformer, util
from io import BytesIO
from openpyxl.styles import PatternFill, Font

# Title and Intro
st.set_page_config(page_title="SKV Standards and Tender Brief Comparator", layout="wide")
st.title("📄 SKV Standards and Tender Brief Comparator")
st.markdown("""
This tool compares your SKV legal standards with any Tender Brief document using **AI-powered semantic analysis**. 
It identifies clause-level **matches, mismatches, and additional tender items** that are not part of SKV.

Upload your files below to begin:
- `Legal Worksheet - Standard.xlsx` (SKV standard file)
- `Any Tender Topsheet.xlsx` (Client-specific tender file)
""")

# File upload
skv_file = st.file_uploader("Upload SKV Standards Excel File", type="xlsx")
tender_file = st.file_uploader("Upload Tender Topsheet Excel File", type="xlsx")

if skv_file and tender_file:
    with st.spinner("Processing files with semantic engine..."):
        # Load files
        skv_df = pd.read_excel(skv_file)
        tender_df = pd.read_excel(tender_file)

        # Extract SKV clauses
        skv_clauses = skv_df[['Clauses', 'SKV Standard']].dropna()

        # Tender brief with Doc Name and Page Number
        tender_brief = tender_df.iloc[1:, [1, 2, 3]]
        tender_brief.columns = ['Tender Brief', 'Value', 'Doc Name and Page Number']
        tender_brief = tender_brief.dropna()

        # Load SBERT model
        model = SentenceTransformer('all-MiniLM-L6-v2')
        skv_embeddings = model.encode(skv_clauses['Clauses'].tolist(), convert_to_tensor=True)
        tender_embeddings = model.encode(tender_brief['Tender Brief'].tolist(), convert_to_tensor=True)

        # Semantic matching
        results = []
        matched_tender_indices = set()

        for i, skv_clause in skv_clauses.iterrows():
            cosine_scores = util.cos_sim(skv_embeddings[i], tender_embeddings)[0]
            best_match_idx = int(np.argmax(cosine_scores))
            score = float(cosine_scores[best_match_idx])

            tender_row = tender_brief.iloc[best_match_idx]
            matched_tender_indices.add(best_match_idx)

            if score > 0.85:
                inference = "✅ Match"
                fill_color = "C6EFCE"  # green
            elif 0.6 < score <= 0.85:
                inference = "🟡 Needs Clarification"
                fill_color = "FFF2CC"  # yellow
            else:
                inference = "❌ Conflict or Not Found"
                fill_color = "F4CCCC"  # red

            results.append({
                "SKV Standards": f"{skv_clause['Clauses']}: {skv_clause['SKV Standard']}",
                "Tender Brief": f"{tender_row['Tender Brief']}: {tender_row['Value']}",
                "Inference": inference,
                "Doc Name and Page Number": tender_row['Doc Name and Page Number'],
                "Fill Color": fill_color
            })

        comparison_df = pd.DataFrame(results)
        fill_colors = comparison_df.pop("Fill Color")

        # Identify extra tender fields
        extra_rows = []
        for i, row in tender_brief.iterrows():
            if i not in matched_tender_indices:
                extra_rows.append({
                    "Tender Brief Extra Field": row['Tender Brief'],
                    "Value": row['Value'],
                    "Doc Name and Page Number": row['Doc Name and Page Number'],
                    "Comment": "Not part of SKV Standards"
                })

        extra_df = pd.DataFrame(extra_rows)

        # Display tables
        def highlight_rows(row):
            color = fill_colors[row.name]
            css = f'background-color: #{color}; color: black;'
            return [css] * len(row)

        st.subheader("🧾 SKV vs Tender Comparison")
        st.dataframe(comparison_df.style.apply(highlight_rows, axis=1), use_container_width=True)

        st.subheader("🟨 Extra Tender Fields (Not in SKV)")
        st.dataframe(extra_df.style.applymap(lambda x: 'background-color: #fff3cd; color: black'), use_container_width=True)

        # Save to Excel in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            comparison_df.to_excel(writer, index=False, sheet_name="SKV vs Tender")
            extra_df.to_excel(writer, index=False, sheet_name="Extra Tender Fields")

            wb = writer.book
            ws = writer.sheets["SKV vs Tender"]
            for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3)):
                color = fill_colors.iloc[i]
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                for cell in row:
                    cell.fill = fill
                    cell.font = Font(color="000000")

            ws_extra = writer.sheets["Extra Tender Fields"]
            yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            for row in ws_extra.iter_rows(min_row=2, max_row=ws_extra.max_row, min_col=1, max_col=4):
                for cell in row:
                    cell.fill = yellow_fill
                    cell.font = Font(color="000000")

        output.seek(0)

        # Download button
        st.download_button(
            label="📥 Download Excel Report",
            data=output,
            file_name="SKV_Tender_Comparison_Result.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# Footer
st.markdown("""
---
Made with ❤️ by [adityxrai](https://adityxrai.vercel.app)
""")
